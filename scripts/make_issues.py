# github issuesを一括作成するスクリプト
# 使うときは、config.pyと階層を合わせて、ACCESS_TOKENとREPO_NAMEを設定してください。
# Due Dateが反映されないので、修正必要あり

"""xlsx から GitHub Issue を一括作成するスクリプト

シートの構成 (1 行目はヘッダで読み飛ばす)
A列: Milestone
B列: Labels (カンマ区切りまたは空白区切り)
C列: Issues (タイトル)
D列: Due Date (YYYY-MM-DD 形式など)

使用例:
    python scripts/create_issues_from_xlsx.py

config.py で ACCESS_TOKEN, REPO_NAME を設定しておくこと。
"""

from __future__ import annotations

import datetime
from typing import Iterable, Optional

from github import Github
from github.GithubException import GithubException
from openpyxl import load_workbook

from config import ACCESS_TOKEN, REPO_NAME

XLSX_PATH = "docs/document.xlsx"


def parse_labels(raw: Optional[str]) -> list[str]:
    """ラベル文字列をリストに変換する。
    カンマまたは空白で区切る。空なら空リスト。
    """
    if not raw:
        return []
    # split by comma or whitespace
    parts = [p.strip() for p in raw.replace("，", ",").split(",")]
    labels: list[str] = []
    for part in parts:
        if not part:
            continue
        labels.extend(part.split())
    return [l for l in labels if l]


def find_or_create_milestone(repo, title: str, due_on: Optional[datetime.date]) -> Optional[object]:
    if not title:
        return None

    existing = {m.title: m for m in repo.get_milestones(state="all")}
    if title in existing:
        return existing[title]

    # create a new milestone
    try:
        m = repo.create_milestone(title=title, due_on=due_on)
        print(f"新規マイルストーン作成: {title} (due {due_on})")
        return m
    except GithubException as e:
        print(f"マイルストーン作成エラー '{title}': {e}")
        return None


def load_rows(path: str) -> Iterable[list]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        # 偶然すべてNoneなら読み取り終了
        if all(cell is None for cell in row):
            continue
        yield list(row)


def main():
    g = Github(ACCESS_TOKEN)
    repo = g.get_repo(REPO_NAME)

    for row in load_rows(XLSX_PATH):
        milestone_title = row[0] or ""
        labels_raw = row[1] or ""
        issue_title = row[2] or ""
        due_date_cell = row[3]

        if not issue_title:
            print("タイトルが指定されていない行をスキップ")
            continue

        due_on = None
        if due_date_cell:
            if isinstance(due_date_cell, datetime.datetime):
                due_on = due_date_cell.date()
            elif isinstance(due_date_cell, datetime.date):
                due_on = due_date_cell
            else:
                # try parse string
                try:
                    due_on = datetime.datetime.strptime(str(due_date_cell), "%Y-%m-%d").date()
                except ValueError:
                    print(f"日付形式不正: {due_date_cell}")

        milestone_obj = find_or_create_milestone(repo, milestone_title, due_on)
        labels = parse_labels(labels_raw)

        try:
            repo.create_issue(title=issue_title, body="", milestone=milestone_obj, labels=labels)
            print(f"Issue作成: {issue_title}")
        except GithubException as e:
            print(f"Issue作成エラー '{issue_title}': {e}")


if __name__ == "__main__":
    main()
